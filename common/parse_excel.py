# !/usr/bin/python3
# -*- coding: utf-8 -*-
import os
from openpyxl import load_workbook


class ParseExcel(object):
    def __init__(self, excelPath, sheetName):
        print(excelPath, sheetName)
        self.wb = load_workbook(excelPath)
        self.sheet = self.wb[sheetName]
        self.maxRowNum = self.sheet.max_row

    def getDatasFromSheet(self):
        dataList = []
        for line in self.sheet.rows:
            tmpList = []
            tmpList.append(line[0].value)
            tmpList.append(line[1].value)
            tmpList.append(line[2].value)
            tmpList.append(line[3].value)
            tmpList.append(line[4].value)
            tmpList.append(line[5].value)
            tmpList.append(line[6].value)
            tmpList.append(line[7].value)
            dataList.append(tmpList)
        return dataList[1:]


def order_api_v8_test_data():
    """
    从外部获取参数数据
    :return:
    """
    path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'test_data\cn_pord')
    excelPath = os.path.join(path, 'test_order_profess.xlsx')
    print(excelPath)
    sheetName = 'order_api_v8_test_data'
    return ParseExcel(excelPath, sheetName)

if __name__ == '__main__':
    print(order_api_v8_test_data())